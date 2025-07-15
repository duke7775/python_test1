#说明：读写Excel文件
import openpyxl

#创建Excel文件
def create_excel_file(filename)-> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("NewSheet")
    sheet["A1"] = 30
    sheet["A2"] = 40
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    workbook.save(filename)

#读取Excel文件
def read_excel_file(filename='test.xlsx')-> None:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["NewSheet"]
    value_a1 = sheet["A1"].value
    value_a2 = sheet["A2"].value
    print(f"A1 值：{value_a1}, A2 值：{value_a2}")
    sheet["A3"] = value_a1 + value_a2
    workbook.save(filename)

def main():
    filename = 'd:/github/test.xlsx'
    create_excel_file(filename)
    read_excel_file(filename)

if __name__ == "__main__":
    main()   
